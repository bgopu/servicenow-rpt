"""
ServiceNow ↔ ADF Correlation

Maps ServiceNow JOBFAILURE incidents to their root ADF pipeline run + error.
Job → pipeline mapping is loaded from Autosys_Job_details_P01.xlsx (SummaryProd sheet).

Scope: ALL domains (Customer, Supplier, Finance, Item, Reference, Worker).

Usage:
    python correlate_incidents_adf.py                                       # last 7 days
    python correlate_incidents_adf.py 14                                    # last 14 days
    python correlate_incidents_adf.py 14 C:\\path\\to\\Incidents_list.csv   # custom CSV
"""

import json
import logging
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Force UTF-8 output so special chars don't break on Windows terminals
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd

logging.getLogger("azure.identity").setLevel(logging.ERROR)
logging.getLogger("azure.core").setLevel(logging.ERROR)

from adf_fetcher import get_credential, get_failed_runs, get_activity_errors, _extract_root_cause

# Project root is one level above src/
_ROOT = Path(__file__).parent.parent

# All domain ADF config files
DOMAIN_CONFIGS = [
    _ROOT / "config" / "config_customer_adf.json",
    _ROOT / "config" / "config_supplier_adf.json",
    _ROOT / "config" / "config_finance_adf.json",
    _ROOT / "config" / "config_item_adf.json",
    _ROOT / "config" / "config_reference_adf.json",
    _ROOT / "config" / "config_worker_adf.json",
]
_default_incidents_csv = _ROOT / "reports" / "Incidents_list.csv"
INCIDENTS_CSV = Path(sys.argv[2]) if len(sys.argv) > 2 and not sys.argv[2].startswith("--") else _default_incidents_csv

MAPPING_XLSX = _ROOT / "Autosys_Job_details_P01.xlsx"
SUBJECT_AREA_FILTER = ""  # All domains


# ── Load job → pipeline + RG mapping from Excel ──────────────────────────────
def load_job_pipeline_map(
    xlsx_path: Path = MAPPING_XLSX,
    subject_area: str = SUBJECT_AREA_FILTER,
) -> dict[str, dict]:
    """
    Read Autosys CMD → {pipeline, rg} from the Excel sheet, filtered to a subject area.
    Returns a dict keyed by lowercase job name for case-insensitive exact lookup.
    Each value: {"pipeline": str, "rg": str}
    """
    df = pd.read_excel(xlsx_path, sheet_name="SummaryProd")
    if subject_area:
        df = df[df["Subject Area"].str.contains(subject_area, na=False, case=False)]
    mapping = (
        df[["Autosys CMD", "Pipeline Name", "RG"]]
        .dropna(subset=["Autosys CMD", "Pipeline Name", "RG"])
        .drop_duplicates(subset=["Autosys CMD"])
    )
    result = {
        str(row["Autosys CMD"]).strip().lower(): {
            "pipeline": str(row["Pipeline Name"]).strip(),
            "rg":       str(row["RG"]).strip(),
        }
        for _, row in mapping.iterrows()
    }
    label = subject_area if subject_area else "all domains"
    print(f"[Mapping] Loaded {len(result)} job→pipeline entries ({label}).")
    return result


# Built once at import time
JOB_TO_PIPELINE: dict[str, dict] = load_job_pipeline_map()

def extract_job_name(short_description: str) -> str:
    """Pull the Autosys job name from the incident short description."""
    # Format: '<jobname>^P01 JOBFAILURE <server>'
    if isinstance(short_description, str) and "JOBFAILURE" in short_description:
        return short_description.split("^")[0].strip().lower()
    return ""


# ── RG overrides: fix any incorrect RG values from the Excel ─────────────────
# Maps job name → correct resource group when the Excel has a wrong RG entry
RG_OVERRIDES: dict[str, str] = {
    # azmcdcpcus101 and azmcdcusingssidecar belong to cp-rg-mdcus-prod-sj per Excel
    # (no overrides needed — Excel is correct; access needed on cp-rg-mdcus-prod-sj)
}


def job_to_pipeline(job_name: str) -> dict | None:
    """
    Map an Autosys job name to {pipeline, rg} using the Excel mapping.
    Returns None if the job is outside the current domain scope (POC: Customer only).
    """
    entry = JOB_TO_PIPELINE.get(job_name.lower())
    if entry and job_name.lower() in RG_OVERRIDES:
        entry = {**entry, "rg": RG_OVERRIDES[job_name.lower()]}
    return entry


def find_best_adf_match(adf_runs: list, pipeline_name: str, incident_time) -> dict | None:
    """
    Find the ADF run for a given pipeline that is closest in time to the incident.
    Returns None if no run found within 4 hours.
    """
    if incident_time.tzinfo is None:
        incident_time = incident_time.replace(tzinfo=timezone.utc)

    candidates = [r for r in adf_runs if r.pipeline_name == pipeline_name]
    if not candidates:
        return None

    best = None
    best_delta = timedelta(hours=12)  # max window — ADF may run hours after Autosys job

    for run in candidates:
        run_start = run.run_start
        if run_start is None:
            continue
        if run_start.tzinfo is None:
            run_start = run_start.replace(tzinfo=timezone.utc)

        # Incident opens after or during the ADF run
        delta = abs(incident_time - run_start)
        if delta < best_delta:
            best_delta = delta
            best = run

    return best


def _append_to_excel(xlsx_path: Path, rows: list):
    """Append rows to the persistent Excel file, skipping duplicates by (incident, adf_run_id, failed_activities)."""
    if not rows:
        return
    # We need EXCEL_HEADERS — define the minimal set needed here inline
    from openpyxl.utils import get_column_letter as _gcl
    EXCEL_HEADERS = [
        ("Incident",           "incident",          14),
        ("Caller",             "caller_id",          20),
        ("Short Description",  "short_description",  45),
        ("Business Service",   "business_service",   25),
        ("Priority",           "priority",           10),
        ("State",              "state",              12),
        ("Assignment Group",   "assignment_group",   28),
        ("Assigned To",        "assigned_to",        22),
        ("Opened (UTC)",       "opened_at",          18),
        ("Breach Reason",      "u_breach_reason",    22),
        ("Job Name",           "job_name",           28),
        ("Pipeline",           "pipeline",           45),
        ("RG",                 "rg",                 28),
        ("ADF",                "adf",                28),
        ("ADF Run ID",         "adf_run_id",         38),
        ("Run Start (UTC)",    "run_start",          18),
        ("Run End (UTC)",      "run_end",            18),
        ("Failed Activity",    "failed_activities",  60),
        ("Error Message",      "root_cause",         80),
        ("Failed Step Input",  "failed_step_input",  70),
        ("Failed Step Output", "failed_step_output", 70),
        ("Status",             "status",             22),
    ]
    WRAP_FROM = {"Failed Activity", "Error Message", "Failed Step Input", "Failed Step Output", "Short Description"}
    existing_rows: list = []
    existing_keys: set = set()
    if xlsx_path.exists():
        try:
            ex_wb = openpyxl.load_workbook(xlsx_path)
            ex_ws = ex_wb.active
            ex_headers = [cell.value for cell in ex_ws[1]]
            for ex_row in ex_ws.iter_rows(min_row=2, values_only=True):
                rd = dict(zip(ex_headers, ex_row))
                mapped = {f: rd.get(lbl, "") for lbl, f, _ in EXCEL_HEADERS}
                existing_rows.append(mapped)
                key = (str(mapped.get("incident","") or ""), str(mapped.get("adf_run_id","") or ""), str(mapped.get("failed_activities","") or ""))
                existing_keys.add(key)
        except Exception:
            pass
    new_rows = []
    for row in rows:
        key = (str(row.get("incident","") or ""), str(row.get("adf_run_id","") or ""), str(row.get("failed_activities","") or ""))
        if key not in existing_keys:
            new_rows.append(row)
            existing_keys.add(key)
    if not new_rows:
        return
    all_rows = existing_rows + new_rows
    xlsx_path.parent.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Domain Job Failures"
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (label, _, width) in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[_gcl(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    # Group rows by incident
    from collections import OrderedDict as _OD
    inc_groups: dict = _OD()
    for row in all_rows:
        inc = str(row.get("incident", "") or "")
        inc_groups.setdefault(inc, []).append(row)
    summary_fill_even = PatternFill("solid", fgColor="BDD7EE")
    summary_fill_odd  = PatternFill("solid", fgColor="DEEAF1")
    detail_fill       = PatternFill("solid", fgColor="F2F2F2")
    summary_font_bold = Font(bold=True)
    ws.sheet_properties.outlinePr.summaryBelow = False
    row_idx = 2
    for inc_num, (inc, rows) in enumerate(inc_groups.items()):
        s_fill = summary_fill_even if inc_num % 2 == 0 else summary_fill_odd
        summary = rows[0]
        for col_idx, (label, field, _) in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=summary.get(field, ""))
            cell.fill = s_fill
            cell.font = summary_font_bold
            cell.alignment = Alignment(vertical="top", wrap_text=(label in WRAP_FROM))
        ws.row_dimensions[row_idx].outline_level = 0
        row_idx += 1
        if len(rows) > 1:
            for detail_row in rows[1:]:
                for col_idx, (label, field, _) in enumerate(EXCEL_HEADERS, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=detail_row.get(field, ""))
                    cell.fill = detail_fill
                    cell.alignment = Alignment(vertical="top", wrap_text=(label in WRAP_FROM))
                ws.row_dimensions[row_idx].outline_level = 1
                ws.row_dimensions[row_idx].hidden = True
                row_idx += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_gcl(len(EXCEL_HEADERS))}1"
    wb.save(xlsx_path)


def _emit_adf_json_from_excel(xlsx_path: Path):
    """Re-build adf_errors.json from the existing Excel when no new incidents need processing."""
    _snow_reports = Path(__file__).parent.parent / "reports"
    if not _snow_reports.exists() or not xlsx_path.exists():
        return
    try:
        xb = openpyxl.load_workbook(xlsx_path)
        xw = xb.active
        hdrs = [cell.value for cell in xw[1]]
        rows = [dict(zip(hdrs, r)) for r in xw.iter_rows(min_row=2, values_only=True)]
        # Map display headers → field names (reuse EXCEL_HEADERS labels)
        label_to_field = {
            "Incident": "incident", "ADF Run ID": "adf_run_id",
            "ADF Pipeline": "pipeline", "Status": "status",
            "Error Message": "root_cause", "Failed Activity": "failed_activities",
            "Failed Step Input": "failed_step_input", "Failed Step Output": "failed_step_output",
        }
        adf_by_inc: dict = {}
        seen_acts: dict = {}
        for rd in rows:
            inc = str(rd.get("Incident", "") or "").strip()
            if not inc:
                continue
            if inc not in adf_by_inc:
                adf_by_inc[inc] = {
                    "pipeline":   str(rd.get("Pipeline", "") or ""),
                    "status":     str(rd.get("Status", "") or ""),
                    "root_cause": str(rd.get("Error Message", "") or ""),
                    "activities": [],
                }
                seen_acts[inc] = set()
            act = str(rd.get("Failed Activity", "") or "")
            if act and act not in seen_acts[inc]:
                seen_acts[inc].add(act)
                adf_by_inc[inc]["activities"].append({
                    "activity": act,
                    "message":  str(rd.get("Error Message", "") or ""),
                    "input":    str(rd.get("Failed Step Input", "") or ""),
                    "output":   str(rd.get("Failed Step Output", "") or ""),
                })
        _adf_path = _snow_reports / "adf_errors.json"
        with open(_adf_path, "w", encoding="utf-8") as f:
            json.dump(adf_by_inc, f, indent=2, ensure_ascii=False)
        print(f"💾 ADF errors JSON refreshed → {_adf_path}  ({len(adf_by_inc)} incidents)")
    except Exception as e:
        print(f"[warn] Could not refresh adf_errors.json: {e}")


def run_correlation(days_back: int = 7):
    # Merge all ADF instances from all domain config files
    adf_instances: list[dict] = []
    for cfg_path in DOMAIN_CONFIGS:
        if not cfg_path.exists():
            print(f"  [Config] {cfg_path.name} not found — skipping")
            continue
        try:
            cfg_data = json.loads(cfg_path.read_text())
            instances = cfg_data.get("adf_instances", [])
            # Add env label if missing
            for inst in instances:
                if "env" not in inst:
                    inst["env"] = inst.get("resource_group", "unknown")
            adf_instances.extend(instances)
        except Exception as e:
            print(f"  [Config] Failed to load {cfg_path.name}: {e}")
    print(f"[Config] Loaded {len(adf_instances)} ADF instance(s) across all domains.")

    # ── Load ServiceNow incidents ──────────────────────────────────────────────
    df = pd.read_csv(INCIDENTS_CSV)
    df["opened_at"] = pd.to_datetime(df["opened_at"], format="ISO8601", errors="coerce")

    year_start = datetime(datetime.now().year, 1, 1)  # Jan 1 of current year
    cutoff_adf = datetime.now() - timedelta(days=days_back)  # window for ADF lookup

    # ALL JOBFAILURE incidents year-to-date → go into Excel
    job_fail_ytd = df[
        df["short_description"].str.contains("JOBFAILURE", na=False)
        & (df["opened_at"] >= year_start)
    ].sort_values("opened_at", ascending=False)

    # Only the recent slice gets ADF correlation
    job_fail_recent = job_fail_ytd[job_fail_ytd["opened_at"] >= cutoff_adf]

    if job_fail_ytd.empty:
        print(f"No JOBFAILURE incidents in {datetime.now().year}.")
        return

    print(f"Found {len(job_fail_ytd)} JOBFAILURE incident(s) YTD ({len(job_fail_recent)} in last {days_back} days).")

    # ── Skip incidents already processed in previous runs ────────────────────
    _logs_dir = Path(__file__).parent.parent / "logs"
    _out_xlsx = _logs_dir / "incidents_to_adf_job_failures.xlsx"
    already_processed: set[str] = set()
    if _out_xlsx.exists():
        try:
            _ex_wb = openpyxl.load_workbook(_out_xlsx)
            _ex_ws = _ex_wb.active
            _ex_headers = [cell.value for cell in _ex_ws[1]]
            if "Incident" in _ex_headers:
                _inc_col = _ex_headers.index("Incident")
                for _ex_row in _ex_ws.iter_rows(min_row=2, values_only=True):
                    _v = _ex_row[_inc_col]
                    if _v:
                        already_processed.add(str(_v).strip())
        except Exception:
            pass

    # Older YTD incidents not yet in Excel → add them with no ADF data
    older_ytd = job_fail_ytd[job_fail_ytd["opened_at"] < cutoff_adf]
    older_new = older_ytd[~older_ytd["number"].astype(str).isin(already_processed)]
    older_rows: list[dict] = []
    for _, inc in older_new.iterrows():
        job_name = extract_job_name(inc["short_description"])
        older_rows.append({
            "incident":          inc["number"],
            "caller_id":         inc.get("caller_id", ""),
            "short_description": inc.get("short_description", ""),
            "business_service":  inc.get("business_service", ""),
            "priority":          inc.get("priority", ""),
            "state":             inc.get("state", ""),
            "assignment_group":  inc.get("assignment_group", ""),
            "assigned_to":       inc.get("assigned_to", ""),
            "opened_at":         inc["opened_at"].strftime("%Y-%m-%d %H:%M") if hasattr(inc["opened_at"], "strftime") else str(inc["opened_at"]),
            "job_name":          job_name,
            "pipeline":          "",
            "failed_activities": "",
            "failed_step_input": "",
            "failed_step_output":"",
            "root_cause":        "",
            "rg":                "",
            "adf":               "",
            "adf_run_id":        "",
            "run_start":         "",
            "run_end":           "",
            "status":            "Outside ADF window",
        })
    if older_rows:
        print(f"  → {len(older_rows)} older YTD incident(s) added to Excel (no ADF lookup).")

    # New incidents within the ADF window
    new_job_fail = job_fail_recent[~job_fail_recent["number"].astype(str).isin(already_processed)]
    skipped_count = len(job_fail_recent) - len(new_job_fail)
    if skipped_count:
        print(f"  → {skipped_count} recent incident(s) already processed — skipping ADF lookup.")

    if new_job_fail.empty:
        print("No new recent incidents — no ADF queries needed.\n")
        # Still append older_rows and refresh JSON
        _csv_rows_extra = older_rows
        if _csv_rows_extra:
            _append_to_excel(_out_xlsx, _csv_rows_extra)
        _emit_adf_json_from_excel(_out_xlsx)
        return

    print(f"  → {len(new_job_fail)} new incident(s) to correlate with ADF.\n")
    job_fail = new_job_fail

    # ── Determine which ADF instances are actually needed ─────────────────────
    # Collect exact RGs from incident→job→pipeline mapping
    needed_rgs: set[str] = set()
    for _, inc in job_fail.iterrows():
        jn = extract_job_name(inc["short_description"])
        info = job_to_pipeline(jn)
        if info:
            needed_rgs.add(info["rg"].lower())

    # Match only the exact ADF instance for each needed RG
    target_instances = [
        inst for inst in adf_instances
        if inst["resource_group"].lower() in needed_rgs
    ]

    if not target_instances:
        print("⚠ No ADF instances matched the incident RGs — falling back to all instances.")
        target_instances = adf_instances

    rgs_label = ", ".join(sorted(needed_rgs)) or "all"
    print(f"RGs in scope      : {rgs_label}")
    print(f"ADF instances     : {len(target_instances)} of {len(adf_instances)} (exact match only)")
    print(f"Fetching ADF failed runs ...")
    all_adf_runs: list = []
    # Map run_id → (client, cfg) for later activity drill-down
    run_client_map: dict[str, tuple] = {}
    for inst in target_instances:
        cfg = {
            "subscription_id": inst["subscription_id"],
            "resource_group":  inst["resource_group"],
            "adf_name":        inst["adf_name"],
        }
        label = f"{inst['adf_name']} ({inst['env']})"
        try:
            runs, c = get_failed_runs(cfg, days_back + 1)
            for r in runs:
                run_client_map[r.run_id] = (c, cfg)
            all_adf_runs.extend(runs)
            print(f"  {label}: {len(runs)} failed run(s)")
        except Exception as e:
            print(f"  {label}: ⚠ Could not fetch runs — {e}")
    print(f"Total: {len(all_adf_runs)} failed ADF run(s).\n")

    # Cache activity errors per run_id to avoid duplicate API calls
    error_cache: dict[str, list] = {}

    print("=" * 80)
    print(f"{'INCIDENT':<13}  {'JOB NAME':<30}  {'OPENED (UTC)':<20}")
    print(f"{'ADF PIPELINE':<44}  {'RUN START (UTC)':<22}  STATUS")
    print("=" * 80)

    # Group by ADF run so we only print the error details once per unique match
    seen_run_ids: set[str] = set()
    csv_rows: list[dict] = []

    def make_row(inc, job_name: str) -> dict:
        """Base row with all ServiceNow fields."""
        return {
            "incident":          inc["number"],
            "caller_id":         inc.get("caller_id", ""),
            "short_description": inc.get("short_description", ""),
            "business_service":  inc.get("business_service", ""),
            "priority":          inc.get("priority", ""),
            "state":             inc.get("state", ""),
            "assignment_group":  inc.get("assignment_group", ""),
            "assigned_to":       inc.get("assigned_to", ""),
            "opened_at":         inc["opened_at"].strftime("%Y-%m-%d %H:%M") if hasattr(inc["opened_at"], "strftime") else str(inc["opened_at"]),
            # "u_breach_reason":   inc.get("u_breach_reason", ""),
            "job_name":          job_name,
            "pipeline":          "",
            "failed_activities":  "",
            "failed_step_input":  "",
            "failed_step_output": "",
            "root_cause":         "",
            "rg":                 "",
            "adf":               "",
            "adf_run_id":        "",
            "run_start":         "",
            "run_end":           "",
            
            "status":            "",
        }

    skipped = 0
    for _, inc in job_fail.iterrows():
        job_name = extract_job_name(inc["short_description"])
        job_info = job_to_pipeline(job_name)

        if job_info is None:
            skipped += 1
            row = make_row(inc, job_name)
            row["status"] = "Job not in mapping"
            csv_rows.append(row)
            continue  # no job→pipeline entry found in Autosys_Job_details_P01.xlsx

        pipeline = job_info["pipeline"]
        job_rg   = job_info["rg"]
        inc_time = inc["opened_at"]
        if inc_time.tzinfo is None:
            inc_time = inc_time.replace(tzinfo=timezone.utc)

        adf_run = find_best_adf_match(all_adf_runs, pipeline, inc_time)
        inst_match = next((i for i in adf_instances if i["resource_group"] == job_rg), None)

        row = make_row(inc, job_name)
        row["pipeline"] = pipeline
        row["rg"]       = job_rg
        row["adf"]      = inst_match["adf_name"] if inst_match else "(unknown)"

        print(f"\n{'-'*80}")
        print(f"  INC       : {inc['number']}")
        print(f"  Job Name  : {job_name}")
        print(f"  Opened    : {inc_time.strftime('%Y-%m-%d %H:%M')} UTC")
        print(f"  Pipeline  : {pipeline}")
        print(f"  RG        : {job_rg}")
        print(f"  ADF       : {row['adf']}")

        if adf_run is None:
            print(f"  ADF Match : ⚠ No matching ADF run found within 12 hours")
            row["status"] = "No ADF run matched"
            csv_rows.append(row)
            continue

        run_start = adf_run.run_start
        if run_start and run_start.tzinfo is None:
            run_start = run_start.replace(tzinfo=timezone.utc)

        row["adf_run_id"] = adf_run.run_id
        row["run_start"]  = run_start.strftime("%Y-%m-%d %H:%M") if run_start else ""
        row["run_end"]    = adf_run.run_end.strftime("%Y-%m-%d %H:%M") if adf_run.run_end else ""

        print(f"  ADF Run   : {adf_run.run_id}")
        print(f"  Run Start : {row['run_start']} UTC")
        print(f"  Run End   : {row['run_end']} UTC")

        # Only fetch and print errors once per unique ADF run
        if adf_run.run_id not in seen_run_ids:
            seen_run_ids.add(adf_run.run_id)
            if adf_run.run_id not in error_cache:
                run_cfg_client, run_cfg = run_client_map.get(adf_run.run_id, (None, None))
                if run_cfg_client and run_cfg:
                    error_cache[adf_run.run_id] = get_activity_errors(run_cfg_client, run_cfg, adf_run.run_id)
                else:
                    error_cache[adf_run.run_id] = []
            errors = error_cache[adf_run.run_id]

            root = ""
            if adf_run.message:
                root = _extract_root_cause(str(adf_run.message))
                print(f"  Root Cause: {root}")

            if errors:
                for e in errors:
                    print(f"  Failed At : {e['activity']} [{e['type']}]  ErrCode={e['error_code']}")
                    err_row = dict(row)
                    err_row["failed_activities"]  = f"{e['activity']} [{e['type']}] ErrCode={e['error_code']}"
                    err_row["failed_step_input"]  = e.get("input", "")
                    err_row["failed_step_output"] = e.get("output", "")
                    err_row["root_cause"]         = e.get("message", "") or root
                    err_row["status"]             = "Matched"
                    csv_rows.append(err_row)
            else:
                row["root_cause"] = root
                row["status"]     = "Matched (no activity errors)"
                csv_rows.append(row)
        else:
            print(f"  (Same ADF run as above -- error details already shown)")
            prior = [r for r in csv_rows if r["adf_run_id"] == adf_run.run_id]
            for pr in prior:
                dup = dict(pr)
                dup["incident"]          = row["incident"]
                dup["short_description"] = row["short_description"]
                dup["opened_at"]         = row["opened_at"]
                dup["status"]            = "Matched (same run)"
                csv_rows.append(dup)
            if not prior:
                row["status"] = "Matched (same run)"
                csv_rows.append(row)

    # ── Write Excel output ──────────────────────────────────────────────────
    logs_dir = Path(__file__).parent.parent / "logs"
    logs_dir.mkdir(exist_ok=True)
    EXCEL_HEADERS = [
        ("Incident",           "incident",          14),
        ("Caller",             "caller_id",          20),
        ("Short Description",  "short_description",  45),
        ("Business Service",   "business_service",   25),
        ("Priority",           "priority",           10),
        ("State",              "state",              12),
        ("Assignment Group",   "assignment_group",   28),
        ("Assigned To",        "assigned_to",        22),
        ("Opened (UTC)",       "opened_at",          18),
        ("Breach Reason",      "u_breach_reason",    22),
        ("Job Name",           "job_name",           28),
        ("Pipeline",           "pipeline",           45),
        ("RG",                 "rg",                 28),
        ("ADF",                "adf",                28),
        ("ADF Run ID",         "adf_run_id",         38),
        ("Run Start (UTC)",    "run_start",          18),
        ("Run End (UTC)",      "run_end",            18),
        ("Failed Activity",    "failed_activities",  60),
        ("Error Message",      "root_cause",         80),
        ("Failed Step Input",  "failed_step_input",  70),
        ("Failed Step Output", "failed_step_output", 70),
        ("Status",             "status",             22),
    ]
    WRAP_FROM = {"Failed Activity", "Error Message", "Failed Step Input", "Failed Step Output", "Short Description"}

    if csv_rows or older_rows:
        out_xlsx = logs_dir / "incidents_to_adf_job_failures.xlsx"

        # ── Load existing rows so we only append genuinely new ones ──────────
        existing_rows: list = []
        existing_keys: set = set()   # (incident, adf_run_id, failed_activities)
        if out_xlsx.exists():
            try:
                ex_wb = openpyxl.load_workbook(out_xlsx)
                ex_ws = ex_wb.active
                ex_headers = [cell.value for cell in ex_ws[1]]
                for ex_row in ex_ws.iter_rows(min_row=2, values_only=True):
                    rd = dict(zip(ex_headers, ex_row))
                    mapped = {f: rd.get(lbl, "") for lbl, f, _ in EXCEL_HEADERS}
                    existing_rows.append(mapped)
                    key = (
                        str(mapped.get("incident", "") or ""),
                        str(mapped.get("adf_run_id", "") or ""),
                        str(mapped.get("failed_activities", "") or ""),
                    )
                    existing_keys.add(key)
            except Exception as e:
                print(f"[warn] Could not read existing Excel file: {e} — will overwrite.")
                existing_rows, existing_keys = [], set()

        # Only keep rows whose (incident, run_id, activity) combo is new
        new_rows = []
        for row in (csv_rows + older_rows):
            key = (
                str(row.get("incident", "") or ""),
                str(row.get("adf_run_id", "") or ""),
                str(row.get("failed_activities", "") or ""),
            )
            if key not in existing_keys:
                new_rows.append(row)
                existing_keys.add(key)

        all_rows = existing_rows + new_rows

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Domain Job Failures"

        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(color="FFFFFF", bold=True)
        fill_even   = PatternFill("solid", fgColor="DCE6F1")
        fill_odd    = PatternFill("solid", fgColor="FFFFFF")

        # Header row
        for col_idx, (label, _, width) in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 30

        # ── Group rows by incident: summary row (level 0) + detail rows (level 1, collapsed) ──
        # Collect distinct incidents in order
        from collections import OrderedDict
        incidents_order: list[str] = []
        inc_groups: dict = OrderedDict()
        for row in all_rows:
            inc = str(row.get("incident", "") or "")
            if inc not in inc_groups:
                inc_groups[inc] = []
                incidents_order.append(inc)
            inc_groups[inc].append(row)

        # Summary fills: alternating per incident block
        summary_fill_even = PatternFill("solid", fgColor="BDD7EE")   # soft blue summary
        summary_fill_odd  = PatternFill("solid", fgColor="DEEAF1")   # lighter blue summary
        detail_fill       = PatternFill("solid", fgColor="F2F2F2")   # light grey detail
        summary_font_bold = Font(bold=True)

        # Outline control: summary row ABOVE detail rows
        ws.sheet_properties.outlinePr.summaryBelow = False

        row_idx = 2
        for inc_num, (inc, rows) in enumerate(inc_groups.items()):
            # Pick summary fill based on incident index
            s_fill = summary_fill_even if inc_num % 2 == 0 else summary_fill_odd
            # Summary row: use the first row's data (best root_cause already set)
            summary = rows[0]
            for col_idx, (label, field, _) in enumerate(EXCEL_HEADERS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=summary.get(field, ""))
                cell.fill = s_fill
                cell.font = summary_font_bold
                cell.alignment = Alignment(vertical="top", wrap_text=(label in WRAP_FROM))
            ws.row_dimensions[row_idx].outline_level = 0
            row_idx += 1

            # Detail rows (activities): only if more than one row for this incident
            if len(rows) > 1:
                for detail_row in rows[1:]:
                    for col_idx, (label, field, _) in enumerate(EXCEL_HEADERS, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=detail_row.get(field, ""))
                        cell.fill = detail_fill
                        cell.alignment = Alignment(vertical="top", wrap_text=(label in WRAP_FROM))
                    ws.row_dimensions[row_idx].outline_level = 1
                    ws.row_dimensions[row_idx].hidden = True   # collapsed by default
                    row_idx += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_HEADERS))}1"

        wb.save(out_xlsx)
        n_incidents = len(inc_groups)
        print(f"\nOutput saved → {out_xlsx}  ({len(all_rows)} rows / {n_incidents} incidents, {len(new_rows)} new)")

    # ── Save ADF errors JSON for HTML report enrichment ──────────────────────
    _snow_reports = Path(r"C:\Users\bgopu\servicenow-rpt\reports")
    if _snow_reports.exists():
        _adf_by_inc: dict = {}
        _seen_acts: dict = {}   # inc -> set of activity labels (dedup)
        for _row in csv_rows + older_rows:
            _inc = _row.get("incident")
            if not _inc:
                continue
            if _inc not in _adf_by_inc:
                _adf_by_inc[_inc] = {
                    "pipeline":   _row.get("pipeline", ""),
                    "status":     _row.get("status", ""),
                    "root_cause": _row.get("root_cause", ""),  # fallback for no-activity rows
                    "activities": [],
                }
                _seen_acts[_inc] = set()
            # Keep the best (non-empty) root_cause at incident level
            if not _adf_by_inc[_inc]["root_cause"] and _row.get("root_cause"):
                _adf_by_inc[_inc]["root_cause"] = _row.get("root_cause", "")
            _act_label = _row.get("failed_activities", "")
            if _act_label and _act_label not in _seen_acts[_inc]:
                _seen_acts[_inc].add(_act_label)
                _adf_by_inc[_inc]["activities"].append({
                    "activity": _act_label,
                    "message":  _row.get("root_cause", ""),
                    "input":    _row.get("failed_step_input", ""),
                    "output":   _row.get("failed_step_output", ""),
                })
        _adf_json = dict(_adf_by_inc)
        _adf_path = _snow_reports / "adf_errors.json"
        with open(_adf_path, "w", encoding="utf-8") as _f:
            json.dump(_adf_json, _f, indent=2, ensure_ascii=False)
        print(f"💾 ADF errors JSON saved → {_adf_path}  ({len(_adf_json)} incidents)")

    print(f"\n{'='*80}")
    if skipped:
        print(f"Skipped {skipped} incident(s) with no job→pipeline mapping.")
    print(f"Correlation complete. {len(seen_run_ids)} unique ADF run(s) matched.")


if __name__ == "__main__":
    days = int(sys.argv[1]) if len(sys.argv) > 1 else 7
    run_correlation(days)
